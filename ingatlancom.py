import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import datetime
import time
from PIL import Image

# Define file name
file_name = 'prices.xlsx'

# Check if file exists
if os.path.exists(file_name):
    wb = load_workbook(file_name)
else:
    # Create a workbook and select active worksheet
    wb = Workbook()
    sheet1 = wb.active

    # Create headers
    headers = ['Link', 'Megjegyzes', 'Product ID', 'Location', 'Vármegye, teleprész, utca', 'Hirdető']

    # Write headers to the first row of the worksheet
    sheet1.append(headers)

    # Save the workbook
    wb.save(filename=file_name)

# Load or create worksheets
sheet1 = wb['Sheet'] if 'Sheet' in wb.sheetnames else wb.create_sheet('Sheet')
sheet2 = wb['Sheet2'] if 'Sheet2' in wb.sheetnames else wb.create_sheet('Sheet2')

# Selenium WebDriver for screenshots
options = Options()
options.add_argument("--headless")
options.add_argument("start-maximized")
options.add_argument("disable-infobars")
options.add_argument("--disable-extensions")
driver = webdriver.Firefox(options=options)

# Define headers for requests
user_agent = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.93 Safari/537.36"

# Create dictionary to store previous prices
previous_prices = {}

# Get the current date
current_date = datetime.datetime.now().strftime('%Y.%m.%d')

# Determine the current date's column index
date_column = None

# Determine the product ID's column index
product_id_column = None

# Determine the location's column index
location_column = None

# Determine the address's column index
address_column = None

# Determine the advertiser's column index
advertiser_column = None

# Check if the current date column already exists
for column in sheet1.columns:
    if column[0].value == current_date:
        date_column = column[0].column
        print(f"Using existing column for date: {current_date}")
        break

# Add the current date column if it doesn't exist
if date_column is None:
    date_column = sheet1.max_column + 1
    sheet1.cell(row=1, column=date_column, value=current_date)
    print(f"Created new column for date: {current_date}")

# Check if the product ID column exists
headers = sheet1[1]
for index, header in enumerate(headers):
    if header.value == 'Product ID':
        product_id_column = index + 1
        print("Product ID column already exists.")
        break

# Add the product ID column if it doesn't exist
if product_id_column is None:
    product_id_column = sheet1.max_column + 1
    sheet1.cell(row=1, column=product_id_column, value='Product ID')
    print("Created new Product ID column.")

# Check if the location column exists
for index, header in enumerate(headers):
    if header.value == 'Location':
        location_column = index + 1
        print("Location column already exists.")
        break

# Add the location column if it doesn't exist
if location_column is None:
    location_column = product_id_column + 1
    sheet1.cell(row=1, column=location_column, value='Location')
    print("Created new Location column.")

# Check if the address column exists
for index, header in enumerate(headers):
    if header.value == 'Vármegye, teleprész, utca':
        address_column = index + 1
        print("Vármegye, teleprész, utca column already exists.")
        break

# Add the address column if it doesn't exist
if address_column is None:
    address_column = location_column + 1
    sheet1.cell(row=1, column=address_column, value='Vármegye, teleprész, utca')
    print("Created new Vármegye, teleprész, utca column.")

# Check if the advertiser column exists
for index, header in enumerate(headers):
    if header.value == 'Hirdető':
        advertiser_column = index + 1
        print("Hirdető column already exists.")
        break

# Add the advertiser column if it doesn't exist
if advertiser_column is None:
    advertiser_column = address_column + 1
    sheet1.cell(row=1, column=advertiser_column, value='Hirdető')
    print("Created new Hirdető column.")

# Create screenshots folder if it doesn't exist
screenshots_folder = os.path.join(os.getcwd(), 'screenshots')
if not os.path.exists(screenshots_folder):
    os.makedirs(screenshots_folder)

# Create current date folder within screenshots folder if it doesn't exist
current_date_folder = os.path.join(screenshots_folder, current_date)
if not os.path.exists(current_date_folder):
    os.makedirs(current_date_folder)

for index, row in enumerate(sheet1.iter_rows(min_row=2, values_only=True), start=2):
    url = row[0]

    # Load the webpage
    try:
        response = requests.get(url, headers={"User-Agent": user_agent})
        response.raise_for_status()  # Raise exception if not 200 status
    except requests.HTTPError as http_err:
        print(f'HTTP error occurred: {http_err}')  # Python 3.6
        price = "Nincs"
        sheet1.cell(row=index, column=date_column, value=price)
        continue

    # BeautifulSoup to parse the HTML
    soup = BeautifulSoup(response.text, 'html.parser')

    # Check if page does not exist
    if "A keresett oldal nem található!" in soup.text:
        # Set price as not found
        price = "Nincs"
    else:
        # Find price div
        price_div = soup.find('div', {
            'class': 'listing-property justify-content-around col-12 col-sm col-print d-flex flex-column text-center print-border-end border-sm-end border-1 border-ash fs-7 font-family-secondary'})

        if price_div:
            price = price_div.find('span', {'class': 'fw-bold fs-5 text-nowrap'}).text

            # Convert price string to number
            try:
                if "millió Ft" in price:
                    price = price.replace(" millió Ft", "").replace(",", ".")
                    price = float(price) * 1000000  # Convert to numerical format
                else:
                    price = float(price)
            except ValueError:
                print(f"Could not convert price to number for URL: {url}")
                continue
        else:
            # Set price as not found
            price = "Nincs"

    # Get the previous price for the current URL
    previous_price = previous_prices.get(url)

    # Write price to Excel if changed or the current date column doesn't have a value
    if price != previous_price or sheet1.cell(row=index, column=date_column).value is None:
        sheet1.cell(row=index, column=date_column, value=price)
        previous_prices[url] = price
        print(f"Updated price for URL: {url}")

    # Capture screenshot
    driver.get(url)

    # Wait for the cookie button to be clickable and click it
    try:
        WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.ID, "CybotCookiebotDialogBodyLevelButtonLevelOptinAllowAll"))
        ).click()
    except Exception as e:
        print(f"Could not click on the cookie consent button. Error: {e}")

    total_width = driver.execute_script("return document.body.offsetWidth")
    total_height = driver.execute_script("return document.body.parentNode.scrollHeight")
    driver.set_window_size(total_width, total_height)
    time.sleep(2)  # wait for 2 seconds to allow the page to load completely
    screenshot_filename = f"{url.split('/')[-1]}_{current_date}.png"
    screenshot_path = os.path.join(current_date_folder, screenshot_filename)
    driver.save_screenshot(screenshot_path)
    print(f"Screenshot saved: {screenshot_path}")

    # Convert PNG to JPG
    img = Image.open(screenshot_path).convert("RGB")
    jpg_screenshot_path = screenshot_path.replace('.png', '.jpg')
    img.save(jpg_screenshot_path, "JPEG")
    # Delete the PNG file
    os.remove(screenshot_path)
    print(f"Converted screenshot to JPG: {jpg_screenshot_path}")

    # Write product ID to Excel
    product_id = url.split('/')[-1]
    sheet1.cell(row=index, column=product_id_column, value=product_id)
    print(f"Product ID for URL: {url} is {product_id}")

    # Write location to Excel
    location_data = soup.find('span', {
        'class': 'card-title px-0 fw-bold fs-4 mb-0 font-family-secondary'}).text if soup.find('span', {
        'class': 'card-title px-0 fw-bold fs-4 mb-0 font-family-secondary'}) else "NULL"
    if ',' in location_data:
        location, address = location_data.split(',', 1)
    else:
        location, address = location_data, "NULL"

    sheet1.cell(row=index, column=location_column, value=location.strip())
    print(f"Location for URL: {url} is {location.strip()}")

    sheet1.cell(row=index, column=address_column, value=address.strip())
    print(f"Vármegye, teleprész, utca for URL: {url} is {address.strip()}")

    # Write advertiser to Excel
    advertiser_span = soup.find('span', {'class': 'd-block fw-500 font-family-secondary fs-6 text-onyx'})
    advertiser = advertiser_span.text if advertiser_span else 'Magánszemély'
    sheet1.cell(row=index, column=advertiser_column, value=advertiser)
    print(f"Hirdető for URL: {url} is {advertiser}")

# Save the workbook
wb.save(filename=file_name)

# Close the WebDriver
driver.quit()
